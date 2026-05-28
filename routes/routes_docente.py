from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify
from datetime import datetime
from sqlalchemy.orm import joinedload
from db import db
from models import (
    Colaborador, Estudiante, Curso, Evaluacion, Asistencia,
    Comentario, Bimestre, CarpetaDocente, DocumentoDocente,
    SolicitudReporte, Horario, Nivel, Grado, Seccion, Inscripcion
)
from functools import wraps
import os

docente_bp = Blueprint('docente', __name__, url_prefix='/docente')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Debes iniciar sesión primero', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get('rol') not in roles:
                flash('No tienes permisos para acceder aquí', 'danger')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated
    return decorator

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads')

def obtener_bimestre_actual():
    hoy = datetime.now().date()
    return Bimestre.query.filter(
        Bimestre.fecha_inicio <= hoy,
        Bimestre.fecha_fin >= hoy
    ).first()

@docente_bp.route('/dashboard')
@login_required
@role_required('docente')
def dashboard():
    docente = Colaborador.query.get(session['usuario_id'])
    cursos = Curso.query.options(
        joinedload(Curso.grado_rel), joinedload(Curso.seccion_rel), joinedload(Curso.periodo)
    ).filter_by(docente_id=docente.id, activo=True).all()
    curso_ids = [c.id for c in cursos]
    curso_grado_ids = {c.grado_id for c in cursos if c.grado_id}
    curso_seccion_ids = {c.seccion_id for c in cursos if c.seccion_id}
    total_estudiantes = db.session.query(Inscripcion.alumno_id).filter(Inscripcion.curso_id.in_(curso_ids)).distinct().count() if curso_ids else 0
    grados = Grado.query.filter(Grado.id.in_(curso_grado_ids)).all() if curso_grado_ids else []
    nivel_ids = {g.nivel_id for g in grados}
    niveles = Nivel.query.filter(Nivel.id.in_(nivel_ids)).all() if nivel_ids else []
    secciones = Seccion.query.filter(Seccion.id.in_(curso_seccion_ids)).all() if curso_seccion_ids else []
    bimestres = Bimestre.query.all()
    horarios = Horario.query.options(
        joinedload(Horario.curso), joinedload(Horario.seccion), joinedload(Horario.docente)
    ).filter_by(docente_id=docente.id).all()
    solicitudes = SolicitudReporte.query.filter_by(activa=True).order_by(SolicitudReporte.fecha_maxima).all()
    estudiantes_por_curso = {}
    for c in cursos:
        alumno_ids = db.session.query(Inscripcion.alumno_id).filter(Inscripcion.curso_id == c.id).subquery()
        estudiantes_por_curso[c.id] = Estudiante.query.filter(Estudiante.id.in_(alumno_ids), Estudiante.activo == True).order_by(Estudiante.apellido_paterno).all()
    return render_template('dashboard_docente.html', docente=docente, cursos=cursos,
        total_estudiantes=total_estudiantes, niveles=niveles, grados=grados,
        secciones=secciones, bimestres=bimestres, horarios=horarios,
        solicitudes=solicitudes, estudiantes_por_curso=estudiantes_por_curso)

@docente_bp.route('/cursos/<int:curso_id>/evaluaciones', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def evaluaciones_curso(curso_id):
    from app import obtener_bimestre_actual as get_bim, _calcular_promedio_desde_datos
    curso = Curso.query.get_or_404(curso_id)
    if curso.docente_id != session['usuario_id']:
        flash('No tienes permiso para este curso', 'danger')
        return redirect(url_for('docente.dashboard'))
    bimestre_id = request.args.get('bimestre_id', type=int) or (get_bim().id if get_bim() else None)
    estudiantes = Estudiante.query.filter_by(
        seccion_id=curso.seccion_id, activo=True
    ).order_by(Estudiante.apellido_paterno).all()
    bimestres = Bimestre.query.filter_by(periodo_academico_id=curso.periodo_academico_id).all()

    if request.method == 'POST':
        accion = request.form.get('accion')
        if accion == 'guardar':
            estudiante_id = int(request.form.get('estudiante_id'))
            tipo = request.form.get('tipo')
            calificacion = float(request.form.get('calificacion'))
            if 0 <= calificacion <= 20:
                e = Evaluacion(
                    curso_id=curso_id, estudiante_id=estudiante_id,
                    bimestre_id=bimestre_id, tipo=tipo,
                    calificacion=calificacion,
                    fecha=datetime.now().date(),
                    observaciones=request.form.get('observaciones', '')
                )
                db.session.add(e); db.session.commit()
                flash('Calificación guardada', 'success')
            else:
                flash('La nota debe ser entre 0 y 20', 'danger')
        elif accion == 'guardar_todo':
            estudiantes_ids = request.form.getlist('estudiante_id[]')
            tipos = request.form.getlist('tipo[]')
            calificaciones = request.form.getlist('calificacion[]')
            existing = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
            evals_key = {}
            for ev in existing:
                evals_key[(ev.estudiante_id, ev.tipo)] = ev
            for i in range(len(estudiantes_ids)):
                est_id = int(estudiantes_ids[i])
                tipo = tipos[i]
                cal = calificaciones[i]
                if cal and cal.strip():
                    cal_val = float(cal)
                    if 0 <= cal_val <= 20:
                        key = (est_id, tipo)
                        if key in evals_key:
                            evals_key[key].calificacion = cal_val
                        else:
                            ev = Evaluacion(
                                curso_id=curso_id, estudiante_id=est_id,
                                bimestre_id=bimestre_id, tipo=tipo,
                                calificacion=cal_val, fecha=datetime.now().date()
                            )
                            db.session.add(ev)
            db.session.commit()
            flash('Notas guardadas', 'success')
        elif accion == 'importar':
            import openpyxl
            archivo = request.files.get('archivo')
            if archivo and archivo.filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
                dnis = []
                rows_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    dni = str(row[2]) if row[2] is not None else ''
                    if dni:
                        dnis.append(dni)
                        rows_data.append((dni, row[3], row[4], row[5], row[6], row[7]))
                estudiantes_map = {}
                if dnis:
                    for e in Estudiante.query.filter(Estudiante.dni.in_(dnis)).all():
                        estudiantes_map[e.dni] = e
                existing_evals = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
                evals_key = {}
                for ev in existing_evals:
                    evals_key[(ev.estudiante_id, ev.tipo)] = ev
                for dni, cuaderno, libro, practicas, exposiciones, examen in rows_data:
                    est = estudiantes_map.get(dni)
                    if not est: continue
                    for tipo, nota in [('cuaderno', cuaderno), ('libro', libro), ('practicas', practicas),
                                       ('exposiciones', exposiciones), ('examen', examen)]:
                        if nota is not None:
                            key = (est.id, tipo)
                            if key in evals_key:
                                evals_key[key].calificacion = float(nota)
                            else:
                                ev = Evaluacion(curso_id=curso_id, estudiante_id=est.id,
                                    bimestre_id=bimestre_id, tipo=tipo,
                                    calificacion=float(nota), fecha=datetime.now().date())
                                db.session.add(ev)
                db.session.commit()
                flash('Notas importadas desde Excel', 'success')
        return redirect(url_for('docente.evaluaciones_curso', curso_id=curso_id))

    evaluaciones = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
    evals_por_estudiante = {}
    for ev in evaluaciones:
        if ev.estudiante_id not in evals_por_estudiante:
            evals_por_estudiante[ev.estudiante_id] = {}
        if ev.tipo not in evals_por_estudiante[ev.estudiante_id]:
            evals_por_estudiante[ev.estudiante_id][ev.tipo] = []
        evals_por_estudiante[ev.estudiante_id][ev.tipo].append(float(ev.calificacion))

    return render_template('evaluaciones.html',
        curso=curso, estudiantes=estudiantes,
        bimestre_id=bimestre_id, bimestres=bimestres,
        evals=evals_por_estudiante)

@docente_bp.route('/cursos/<int:curso_id>/asistencia', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def asistencia_curso(curso_id):
    from app import obtener_bimestre_actual as get_bim
    curso = Curso.query.get_or_404(curso_id)
    if curso.docente_id != session['usuario_id']:
        flash('No tienes permiso para este curso', 'danger')
        return redirect(url_for('docente.dashboard'))
    bimestre_id = request.args.get('bimestre_id', type=int) or (get_bim().id if get_bim() else None)
    estudiantes = Estudiante.query.filter_by(
        seccion_id=curso.seccion_id, activo=True
    ).order_by(Estudiante.apellido_paterno).all()

    if request.method == 'POST':
        fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
        existentes = Asistencia.query.filter_by(
            curso_id=curso_id, fecha=fecha, bimestre_id=bimestre_id
        ).all()
        existentes_idx = {a.estudiante_id: a for a in existentes}
        for est in estudiantes:
            estado = request.form.get(f'estado_{est.id}', 'presente')
            if est.id in existentes_idx:
                existentes_idx[est.id].estado = estado
            else:
                a = Asistencia(curso_id=curso_id, estudiante_id=est.id,
                    bimestre_id=bimestre_id, fecha=fecha,
                    estado=estado, marcado_por=session['usuario_id'])
                db.session.add(a)
        db.session.commit()
        flash('Asistencia guardada', 'success')
        return redirect(url_for('docente.asistencia_curso', curso_id=curso_id))

    fecha_seleccionada = request.args.get('fecha', datetime.now().strftime('%Y-%m-%d'))
    asistencias_hoy = {}
    asis = Asistencia.query.filter_by(curso_id=curso_id, fecha=fecha_seleccionada, bimestre_id=bimestre_id).all()
    for a in asis:
        asistencias_hoy[a.estudiante_id] = a.estado
    bimestres = Bimestre.query.filter_by(periodo_academico_id=curso.periodo_academico_id).all()
    return render_template('asistencia.html',
        curso=curso, estudiantes=estudiantes,
        fecha=fecha_seleccionada, asistencias=asistencias_hoy,
        bimestre_id=bimestre_id, bimestres=bimestres)

@docente_bp.route('/comentarios', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def comentarios():
    from app import obtener_bimestre_actual as get_bim
    docente = Colaborador.query.get(session['usuario_id'])
    cursos = Curso.query.options(joinedload(Curso.grado_rel), joinedload(Curso.seccion_rel)).filter_by(docente_id=docente.id).all()
    if request.method == 'POST':
        c = Comentario(
            docente_id=docente.id,
            estudiante_id=int(request.form.get('estudiante_id')),
            tipo=request.form.get('tipo'),
            contenido=request.form.get('contenido'),
            bimestre_id=int(request.form.get('bimestre_id')) if request.form.get('bimestre_id') else None
        )
        db.session.add(c); db.session.commit()
        flash('Comentario enviado', 'success')
        return redirect(url_for('docente.comentarios'))
    comentarios = Comentario.query.filter_by(docente_id=docente.id).order_by(Comentario.fecha_creacion.desc()).all()
    return render_template('comentarios_docente.html', cursos=cursos, comentarios=comentarios, bimestres=Bimestre.query.all())

@docente_bp.route('/documentos', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def documentos():
    from werkzeug.utils import secure_filename
    docente = Colaborador.query.get(session['usuario_id'])
    if request.method == 'POST':
        archivo = request.files.get('archivo')
        if not archivo or not archivo.filename:
            flash('No se seleccionó ningún archivo', 'danger')
            return redirect(url_for('docente.documentos'))
        ALLOWED_EXT = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'jpg', 'jpeg', 'png', 'gif', 'txt'}
        ext = archivo.filename.rsplit('.', 1)[1].lower() if '.' in archivo.filename else ''
        if ext not in ALLOWED_EXT:
            flash('Tipo de archivo no permitido. Extensiones: pdf, doc, docx, xls, xlsx, jpg, png, gif, txt', 'danger')
            return redirect(url_for('docente.documentos'))
        filename = secure_filename(archivo.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        archivo.save(filepath)
        solicitud_id = request.form.get('solicitud_reporte_id')
        d = DocumentoDocente(
            carpeta_id=int(request.form.get('carpeta_id')) if request.form.get('carpeta_id') else None,
            docente_id=docente.id,
            solicitud_reporte_id=int(solicitud_id) if solicitud_id else None,
            titulo=request.form.get('titulo'),
            descripcion=request.form.get('descripcion', ''),
            archivo_nombre=filename,
            archivo_ruta=f'uploads/{filename}'
        )
        db.session.add(d); db.session.commit()
        flash('Documento subido', 'success')
        return redirect(url_for('docente.documentos'))
    carpetas = CarpetaDocente.query.filter_by(activo=True).all()
    documentos = DocumentoDocente.query.filter_by(docente_id=docente.id).order_by(DocumentoDocente.fecha_subida.desc()).all()
    ahora = datetime.utcnow()
    solicitudes = SolicitudReporte.query.filter_by(activa=True).filter(
        SolicitudReporte.fecha_maxima > ahora
    ).order_by(SolicitudReporte.fecha_maxima).all()
    return render_template('documentos_docente.html', carpetas=carpetas, documentos=documentos, solicitudes=solicitudes, ahora=ahora)

@docente_bp.route('/api/solicitudes')
@login_required
@role_required('docente')
def api_solicitudes():
    ahora = datetime.utcnow()
    solicitudes = SolicitudReporte.query.filter_by(activa=True).filter(
        SolicitudReporte.fecha_maxima > ahora
    ).order_by(SolicitudReporte.fecha_maxima).all()
    return jsonify([{
        'id': s.id, 'titulo': s.titulo, 'descripcion': s.descripcion,
        'fecha_maxima': s.fecha_maxima.isoformat(),
        'activa': s.fecha_maxima > ahora
    } for s in solicitudes])

@docente_bp.route('/api/asistencia_estudiantes', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def api_asistencia_estudiantes():
    if request.method == 'POST':
        data = request.get_json()
        seccion_id = data.get('seccion_id')
        fecha_str = data.get('fecha')
        bimestre_id = data.get('bimestre_id')
        estudiantes_data = data.get('estudiantes', [])
        if not seccion_id or not fecha_str or not estudiantes_data:
            return jsonify({'success': False, 'error': 'Faltan datos'}), 400
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        docente = Colaborador.query.get(session['usuario_id'])
        existentes = Asistencia.query.filter_by(fecha=fecha, bimestre_id=bimestre_id).all()
        existentes_idx = {a.estudiante_id: a for a in existentes}
        curso = Curso.query.filter_by(seccion_id=seccion_id, docente_id=docente.id).first()
        if not curso:
            return jsonify({'success': False, 'error': 'No tienes un curso en esta seccion'}), 400
        for ed in estudiantes_data:
            est_id = ed.get('id')
            estado = ed.get('estado', 'presente')
            if est_id in existentes_idx:
                existentes_idx[est_id].estado = estado
            else:
                a = Asistencia(curso_id=curso.id, estudiante_id=est_id,
                    bimestre_id=bimestre_id, fecha=fecha, estado=estado,
                    marcado_por=session['usuario_id'])
                db.session.add(a)
        db.session.commit()
        return jsonify({'success': True})
    seccion_id = request.args.get('seccion_id', type=int)
    if not seccion_id:
        return jsonify([])
    estudiantes = Estudiante.query.filter_by(seccion_id=seccion_id, activo=True).order_by(Estudiante.apellido_paterno).all()
    colores = ['#1565c0','#2e7d32','#e65100','#6a1b9a','#c62828','#00838f']
    return jsonify([{
        'id': e.id,
        'nombre': e.nombre_completo,
        'iniciales': (e.nombres[0] if e.nombres else '') + (e.apellido_paterno[0] if e.apellido_paterno else '')
    } for e in estudiantes])

@docente_bp.route('/horario')
@login_required
@role_required('docente')
def horario():
    docente = Colaborador.query.get(session['usuario_id'])
    horarios = Horario.query.options(
        joinedload(Horario.curso), joinedload(Horario.seccion)
    ).filter_by(docente_id=docente.id).all()
    return render_template('docente_horario.html', horarios=horarios, docente=docente,
        dias=['Lunes','Martes','Miércoles','Jueves','Viernes'])

@docente_bp.route('/api/evaluaciones', methods=['GET', 'POST'])
@login_required
@role_required('docente')
def api_evaluaciones():
    from app import nota_a_letra
    docente = Colaborador.query.get(session['usuario_id'])
    if request.method == 'POST':
        data = request.get_json()
        curso_id = data.get('curso_id')
        bimestre_id = data.get('bimestre_id')
        estudiantes_data = data.get('estudiantes', [])
        curso = Curso.query.get(curso_id)
        if not curso or curso.docente_id != docente.id:
            return jsonify({'success': False, 'error': 'No autorizado'}), 403
        existing = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
        evals_key = {}
        for ev in existing:
            evals_key[(ev.estudiante_id, ev.tipo)] = ev
        for ed in estudiantes_data:
            est_id = ed.get('estudiante_id')
            notas = ed.get('notas', {})
            for tipo in ['cuaderno', 'libro', 'practicas', 'exposiciones', 'examen']:
                val = notas.get(tipo)
                if val is not None:
                    key = (est_id, tipo)
                    if key in evals_key:
                        evals_key[key].calificacion = float(val)
                    else:
                        ev = Evaluacion(curso_id=curso_id, estudiante_id=est_id,
                            bimestre_id=bimestre_id, tipo=tipo,
                            calificacion=float(val), fecha=datetime.now().date())
                        db.session.add(ev)
        db.session.commit()
        return jsonify({'success': True})
    curso_id = request.args.get('curso_id', type=int)
    bimestre_id = request.args.get('bimestre_id', type=int)
    if not curso_id or not bimestre_id:
        return jsonify({'estudiantes': []})
    curso = Curso.query.get(curso_id)
    if not curso or curso.docente_id != docente.id:
        return jsonify({'estudiantes': []})
    estudiantes = Estudiante.query.filter_by(seccion_id=curso.seccion_id, activo=True).order_by(Estudiante.apellido_paterno).all()
    evaluaciones = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
    evals_idx = {}
    for ev in evaluaciones:
        evals_idx.setdefault(ev.estudiante_id, {})[ev.tipo] = float(ev.calificacion)
    result = []
    for e in estudiantes:
        notas = evals_idx.get(e.id, {})
        c = notas.get('cuaderno')
        l = notas.get('libro')
        p = notas.get('practicas')
        ex = notas.get('exposiciones')
        em = notas.get('examen')
        prom = None
        if c is not None and l is not None and p is not None and ex is not None and em is not None:
            prom = c*0.10 + l*0.10 + p*0.20 + ex*0.10 + em*0.50
        result.append({
            'id': e.id, 'nombre': e.nombre_completo,
            'notas': {'cuaderno': c, 'libro': l, 'practicas': p, 'exposiciones': ex, 'examen': em},
            'promedio': round(prom, 2) if prom is not None else None,
            'letra': nota_a_letra(prom) if prom is not None else '-'
        })
    return jsonify({'estudiantes': result})

@docente_bp.route('/descargar_plantilla_notas')
@login_required
@role_required('docente')
def descargar_plantilla_notas():
    curso_id = request.args.get('curso_id', type=int)
    if not curso_id:
        flash('Selecciona un curso', 'warning')
        return redirect(url_for('docente.dashboard'))
    curso = Curso.query.get(curso_id)
    if not curso:
        flash('Curso no encontrado', 'danger')
        return redirect(url_for('docente.dashboard'))
    estudiantes = Estudiante.query.filter_by(seccion_id=curso.seccion_id, activo=True).order_by(Estudiante.apellido_paterno).all()
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Notas'
        ws.append(['N°', 'DNI', 'Estudiante', 'Cuaderno(0-20)', 'Libro(0-20)', 'Practicas(0-20)', 'Exposiciones(0-20)', 'Examen(0-20)'])
        for i, e in enumerate(estudiantes, 1):
            ws.append([i, e.dni, e.nombre_completo, '', '', '', '', ''])
        for col in ['D', 'E', 'F', 'G', 'H']:
            for row in range(2, len(estudiantes) + 2):
                cell = ws[f'{col}{row}']
                cell.number_format = '0.0'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename=plantilla_notas_{curso.codigo}.xlsx'})
    except ImportError:
        flash('openpyxl no instalado', 'danger')
        return redirect(url_for('docente.dashboard'))

@docente_bp.route('/importar_notas_excel', methods=['POST'])
@login_required
@role_required('docente')
def importar_notas_excel():
    curso_id = request.form.get('curso_id', type=int)
    bimestre_id = request.form.get('bimestre_id', type=int)
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename.endswith('.xlsx'):
        return jsonify({'success': False, 'error': 'Archivo .xlsx requerido'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        dnis = []
        rows_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            dni = str(row[1]) if row[1] is not None else ''
            if dni:
                dnis.append(dni)
                rows_data.append((dni, row[3], row[4], row[5], row[6], row[7]))
        estudiantes_map = {}
        if dnis:
            for e in Estudiante.query.filter(Estudiante.dni.in_(dnis)).all():
                estudiantes_map[e.dni] = e
        existing = Evaluacion.query.filter_by(curso_id=curso_id, bimestre_id=bimestre_id).all()
        evals_key = {}
        for ev in existing:
            evals_key[(ev.estudiante_id, ev.tipo)] = ev
        for dni, cuaderno, libro, practicas, exposiciones, examen in rows_data:
            est = estudiantes_map.get(dni)
            if not est:
                continue
            for tipo, nota in [('cuaderno', cuaderno), ('libro', libro), ('practicas', practicas),
                               ('exposiciones', exposiciones), ('examen', examen)]:
                if nota is not None:
                    key = (est.id, tipo)
                    if key in evals_key:
                        evals_key[key].calificacion = float(nota)
                    else:
                        ev = Evaluacion(curso_id=curso_id, estudiante_id=est.id,
                            bimestre_id=bimestre_id, tipo=tipo,
                            calificacion=float(nota), fecha=datetime.now().date())
                        db.session.add(ev)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as ex:
        return jsonify({'success': False, 'error': str(ex)})
